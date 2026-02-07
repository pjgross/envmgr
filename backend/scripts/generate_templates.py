"""
Script to generate Excel import templates for EnvManager.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def create_environment_template():
    """Create environment import template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Environments'
    
    # Headers
    headers = ['Name', 'Type', 'Status', 'Owner Email', 'Tags', 'Description']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    ws.append(headers)
    
    # Style headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col_num)].width = 20
    
    # Add data validation for Type column
    type_validation = DataValidation(type="list", formula1='"on-premise,cloud"', allow_blank=False)
    ws.add_data_validation(type_validation)
    type_validation.add('B2:B1000')
    
    # Add data validation for Status column
    status_validation = DataValidation(
        type="list",
        formula1='"active,inactive,maintenance,decommissioned"',
        allow_blank=False
    )
    ws.add_data_validation(status_validation)
    status_validation.add('C2:C1000')
    
    # Add example rows
    ws.append(['QA-ENV-01', 'on-premise', 'active', 'admin@example.com', 'qa,testing', 'QA environment for testing'])
    ws.append(['STAGING-01', 'cloud', 'active', 'admin@example.com', 'staging,aws', 'Staging environment on AWS'])
    
    # Add instructions sheet
    instructions = wb.create_sheet('Instructions')
    instructions.append(['Field', 'Required', 'Description', 'Example'])
    
    inst_data = [
        ['Name', 'Yes', 'Unique environment name', 'QA-ENV-01'],
        ['Type', 'Yes', 'Environment type: on-premise or cloud', 'cloud'],
        ['Status', 'Yes', 'active, inactive, maintenance, or decommissioned', 'active'],
        ['Owner Email', 'Yes', 'Email of environment owner (must exist in system)', 'admin@example.com'],
        ['Tags', 'No', 'Comma-separated tags for categorization', 'qa,testing,aws'],
        ['Description', 'No', 'Detailed description of the environment', 'Primary QA environment']
    ]
    
    for row in inst_data:
        instructions.append(row)
    
    # Style instructions
    for col in range(1, 5):
        instructions.column_dimensions[get_column_letter(col)].width = 25
        cell = instructions.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    return wb


def create_system_template():
    """Create system import template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Systems'
    
    # Headers
    headers = ['Name', 'Description', 'Owner Email', 'GitHub Repository URL']
    header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    ws.append(headers)
    
    # Style headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col_num)].width = 30
    
    # Add example rows
    ws.append(['Payment Service', 'Handles payment processing', 'dev@example.com', 'https://github.com/org/payment-service'])
    ws.append(['User Management', 'User authentication and authorization', 'dev@example.com', 'https://github.com/org/user-mgmt'])
    
    # Add instructions sheet
    instructions = wb.create_sheet('Instructions')
    instructions.append(['Field', 'Required', 'Description', 'Example'])
    
    inst_data = [
        ['Name', 'Yes', 'Unique system name', 'Payment Service'],
        ['Description', 'No', 'System description', 'Handles all payment processing'],
        ['Owner Email', 'Yes', 'Email of system owner (must exist in system)', 'dev@example.com'],
        ['GitHub Repository URL', 'No', 'GitHub repository for infrastructure discovery', 'https://github.com/org/payment-service']
    ]
    
    for row in inst_data:
        instructions.append(row)
    
    # Style instructions
    for col in range(1, 5):
        instructions.column_dimensions[get_column_letter(col)].width = 30
        cell = instructions.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    return wb


def create_project_template():
    """Create project import template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Projects'
    
    # Headers
    headers = ['Name', 'Description', 'Team Members (comma-separated emails)']
    header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    header_font = Font(color='000000', bold=True)
    
    ws.append(headers)
    
    # Style headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col_num)].width = 40
    
    # Add example rows
    ws.append(['Mobile App Team', 'iOS and Android development team', 'dev1@example.com,dev2@example.com,qa@example.com'])
    ws.append(['Web Team', 'Frontend web development team', 'web1@example.com,web2@example.com'])
    
    # Add instructions sheet
    instructions = wb.create_sheet('Instructions')
    instructions.append(['Field', 'Required', 'Description', 'Example'])
    
    inst_data = [
        ['Name', 'Yes', 'Unique project name', 'Mobile App Team'],
        ['Description', 'No', 'Project description', 'iOS and Android mobile applications'],
        ['Team Members', 'Yes', 'Comma-separated list of team member emails', 'dev1@example.com,dev2@example.com']
    ]
    
    for row in inst_data:
        instructions.append(row)
    
    # Style instructions
    for col in range(1, 5):
        instructions.column_dimensions[get_column_letter(col)].width = 40
        cell = instructions.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    return wb


if __name__ == '__main__':
    import os
    
    # Create templates directory if it doesn't exist
    os.makedirs('templates', exist_ok=True)
    
    # Generate templates
    env_wb = create_environment_template()
    env_wb.save('templates/environment_import_template.xlsx')
    print('✓ Created templates/environment_import_template.xlsx')
    
    sys_wb = create_system_template()
    sys_wb.save('templates/system_import_template.xlsx')
    print('✓ Created templates/system_import_template.xlsx')
    
    proj_wb = create_project_template()
    proj_wb.save('templates/project_import_template.xlsx')
    print('✓ Created templates/project_import_template.xlsx')
    
    print('\nAll Excel templates created successfully!')
