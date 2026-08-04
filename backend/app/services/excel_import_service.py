"""
Excel import service — imports environments and systems from .xlsx files.
"""
from io import BytesIO
from typing import Optional

import openpyxl
from fastapi import HTTPException, status
from pydantic import ValidationError
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models.environment import Environment
from app.db.models.environment_tier import EnvironmentTier
from app.db.models.system import System
from app.db.models.user import User
from app.services import environment_service, system_service


def _find_col(headers: list, name: str, required: bool = True) -> Optional[int]:
    """Return the 0-based column index matching *name* (case-insensitive).

    Raises HTTP 400 when required=True and the column is missing.
    """
    lower = name.lower()
    for i, h in enumerate(headers):
        if h is not None and str(h).strip().lower() == lower:
            return i
    if required:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Missing required column: '{name}'",
        )
    return None


async def import_environments(
    db: AsyncSession, file_bytes: bytes, tenant_id: int, imported_by_user_id: int
) -> dict:
    """
    Parse an Excel workbook and import environments.

    Required columns: Name, Type
    Optional columns: Description, Status

    Imported rows get `owner_user_id = imported_by_user_id` — the importer is
    present, acting and identifiable, so recording them as owner is truthful,
    unlike fabricating an owner for a pre-existing row. `expires_at` stays
    null ("no expiry planned"): the spreadsheet has no expiry to offer, and a
    null expiry is no longer treated as a governance gap.

    That owner assignment only holds when the importer is actually a member
    of `tenant_id`. A master admin importing while impersonating a tenant has
    `active_tenant_id == tenant_id` (the impersonated tenant) but their own
    `User.tenant_id` is the system tenant — pointing `owner_user_id` at them
    would fail the same cross-tenant check `_validate_tier_and_owner` applies
    on every other write (`environment_service.py`), and unlike a normal
    validation error that 404 is not caught by the per-row
    `except (ValueError, ValidationError)` below, so it used to fail the
    *entire upload* instead of just this row's owner. Membership is checked
    against the `User` row's own `tenant_id`, not assumed from the caller's
    id, and a non-member imports every row with `owner_user_id=None` instead
    — a deliberate, visible, reportable state (`?governance_gap=true`, and
    listed in `tier_fallbacks` below) rather than a failed import.
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]  # first row = headers

    name_idx = _find_col(headers, "Name", required=True)
    type_idx = _find_col(headers, "Type", required=False)
    desc_idx = _find_col(headers, "Description", required=False)

    # The tier vocabulary is what the tenant's admin configured. A spreadsheet
    # upload resolves against it and never extends it — otherwise anyone who can
    # upload a file can add to a list only an admin is supposed to control.
    tiers = list(
        (
            await db.execute(
                select(EnvironmentTier).where(
                    EnvironmentTier.tenant_id == tenant_id,
                    EnvironmentTier.deleted_at.is_(None),
                )
            )
        )
        .scalars()
        .all()
    )
    if not tiers:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail=(
                "This tenant has no environment tiers configured, so imported "
                "environments cannot be assigned one."
            ),
        )
    tier_by_name = {t.name.strip().lower(): t for t in tiers}
    other_tier = next(
        (t for t in tiers if t.category == "other"), tier_by_name.get("other")
    )

    # See the module docstring above: only set an owner who actually belongs
    # to the target tenant. Checked once against the User row's own
    # `tenant_id`, not assumed from the caller's id — an impersonating
    # master admin's `User.tenant_id` is the system tenant, not `tenant_id`.
    importer = (
        await db.execute(select(User).where(User.id == imported_by_user_id))
    ).scalar_one_or_none()
    importer_is_tenant_member = importer is not None and importer.tenant_id == tenant_id

    errors: list[dict] = []
    tier_fallbacks: list[dict] = []
    created = 0
    skipped = 0

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip completely empty rows
        if all(cell is None for cell in row):
            continue

        name = row[name_idx] if name_idx is not None and name_idx < len(row) else None
        if not name or str(name).strip() == "":
            errors.append({"row": row_num, "field": "Name", "message": "Name is required"})
            continue

        name_str = str(name).strip()

        env_type = (
            str(row[type_idx]).strip()
            if type_idx is not None and type_idx < len(row) and row[type_idx] is not None
            else ""
        )
        # A blank or unrecognised type falls back to Other rather than minting a
        # tier for it. Recorded in `tier_fallbacks` below (once we know the row
        # will actually be created) so a mistyped Type column is reported, not
        # silently filed under Other.
        matched_tier = tier_by_name.get(env_type.strip().lower())
        tier = matched_tier or other_tier
        if tier is None:
            errors.append({
                "row": row_num,
                "field": "Type",
                "message": (
                    f"Unknown environment type '{env_type}' and this tenant has "
                    "no 'Other' tier to fall back to"
                ),
            })
            continue
        description = (
            str(row[desc_idx]).strip()
            if desc_idx is not None and desc_idx < len(row) and row[desc_idx] is not None
            else None
        )

        # Check if env already exists (skip if so)
        existing_result = await db.execute(
            select(Environment).where(
                Environment.name == name_str,
                Environment.tenant_id == tenant_id,
                Environment.deleted_at.is_(None),
            )
        )
        if existing_result.scalar_one_or_none():
            skipped += 1
            continue

        try:
            # Owner is the importing admin (present, acting, identifiable — see
            # the module docstring above) UNLESS they aren't actually a member
            # of this tenant, in which case the row imports unowned rather
            # than failing the whole upload. No expiry either way, since the
            # spreadsheet has none to offer and null now means "no expiry
            # planned" rather than a gap.
            owner_user_id = imported_by_user_id if importer_is_tenant_member else None
            await environment_service.create_environment_record(
                db,
                tenant_id,
                name=name_str,
                description=description or None,
                tier_id=tier.id,
                owner_user_id=owner_user_id,
            )
            created += 1
            if matched_tier is None:
                tier_fallbacks.append({
                    "row": row_num,
                    "field": "Type",
                    "message": (
                        f"Unrecognised type '{env_type}' — filed under the "
                        "tenant's Other tier"
                        if env_type
                        else "No type specified — filed under the tenant's Other tier"
                    ),
                })
            if owner_user_id is None:
                tier_fallbacks.append({
                    "row": row_num,
                    "field": "Owner",
                    "message": (
                        "Importing user is not a member of this tenant — "
                        "imported without an owner"
                    ),
                })
        except (ValueError, ValidationError) as e:
            errors.append({"row": row_num, "field": "general", "message": str(e)})

    return {
        "created": created,
        "skipped": skipped,
        "errors": errors,
        "tier_fallbacks": tier_fallbacks,
    }


async def import_systems(
    db: AsyncSession, file_bytes: bytes, tenant_id: int
) -> dict:
    """
    Parse an Excel workbook and import systems.

    Required columns: Name
    Optional columns: Description, GitHub URL
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]  # first row = headers

    name_idx = _find_col(headers, "Name", required=True)
    desc_idx = _find_col(headers, "Description", required=False)
    github_idx = _find_col(headers, "GitHub URL", required=False)

    errors: list[dict] = []
    created = 0
    skipped = 0

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip completely empty rows
        if all(cell is None for cell in row):
            continue

        name = row[name_idx] if name_idx is not None and name_idx < len(row) else None
        if not name or str(name).strip() == "":
            errors.append({"row": row_num, "field": "Name", "message": "Name is required"})
            continue

        name_str = str(name).strip()

        description = (
            str(row[desc_idx]).strip()
            if desc_idx is not None and desc_idx < len(row) and row[desc_idx] is not None
            else None
        )
        github_url = (
            str(row[github_idx]).strip()
            if github_idx is not None and github_idx < len(row) and row[github_idx] is not None
            else None
        )

        # Check if system already exists (skip if so)
        existing_result = await db.execute(
            select(System).where(
                System.name == name_str,
                System.tenant_id == tenant_id,
                System.deleted_at.is_(None),
            )
        )
        if existing_result.scalar_one_or_none():
            skipped += 1
            continue

        try:
            from app.api.v1.schemas.system import SystemCreate
            data = SystemCreate(
                name=name_str,
                description=description or None,
                github_repository_url=github_url or None,
            )
            await system_service.create_system(db, data, tenant_id)
            created += 1
        except (ValueError, ValidationError) as e:
            errors.append({"row": row_num, "field": "general", "message": str(e)})

    return {"created": created, "skipped": skipped, "errors": errors}
