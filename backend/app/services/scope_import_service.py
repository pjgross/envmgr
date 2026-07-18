"""Scope import service — imports release scope items (ReleaseChange) from .xlsx.

Upserts on (release_id, external_key) when an external_key is present; otherwise
inserts. Sets source='spreadsheet'. Per row, builds + validates tenant custom
fields against custom_field_service (mirrors the manual create path). Non-reserved
column headers are matched to custom-field definitions by field_key or label
(case-insensitive); unknown columns are ignored. No per-row events (bulk import).
"""
from io import BytesIO
from typing import Optional

import openpyxl
from fastapi import HTTPException, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.v1.schemas.version import ImportError
from app.db.models.release_change import ReleaseChange
from app.services import custom_field_service

_ENTITY_TYPE = "release_change"
_FIXED_COLUMNS = ["external_key", "title", "description", "change_kind",
                  "external_status", "project_code", "project_name"]
_RESERVED = {c.lower() for c in _FIXED_COLUMNS}


def _header_index(headers: list, name: str, required: bool) -> Optional[int]:
    lower = name.lower()
    for i, h in enumerate(headers):
        if h is not None and str(h).strip().lower() == lower:
            return i
    if required:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"Missing required column: '{name}'")
    return None


def _cell(row: tuple, idx: Optional[int]):
    if idx is None or idx >= len(row):
        return None
    v = row[idx]
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip()
    return s or None


def _coerce(raw, field_type: str):
    """Coerce a spreadsheet cell to the type validate_custom_fields expects."""
    if field_type == "number":
        if isinstance(raw, bool):
            return raw
        try:
            f = float(raw)
            return int(f) if float(f).is_integer() else f
        except (TypeError, ValueError):
            return raw  # leave as-is; validation will flag it
    if field_type == "boolean":
        if isinstance(raw, bool):
            return raw
        s = str(raw).strip().lower()
        if s in ("true", "yes", "1", "y"):
            return True
        if s in ("false", "no", "0", "n"):
            return False
        return raw  # invalid -> validation flags "must be a boolean"
    return str(raw).strip()


async def import_scope(
    db: AsyncSession, file_bytes: bytes, release_id: int, tenant_id: int
) -> dict:
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Could not read spreadsheet")
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = {c: _header_index(headers, c, required=(c in ("title", "change_kind"))) for c in _FIXED_COLUMNS}

    # Map non-reserved headers to custom-field definitions (by field_key or label).
    all_defs = await custom_field_service.list_definitions(db, tenant_id, _ENTITY_TYPE)
    by_key = {d.field_key.lower(): d for d in all_defs}
    by_label = {d.label.lower(): d for d in all_defs}
    custom_cols = []  # (col_idx, definition)
    for i, h in enumerate(headers):
        if h is None:
            continue
        key = str(h).strip()
        if key.lower() in _RESERVED:
            continue
        defn = by_key.get(key.lower()) or by_label.get(key.lower())
        if defn is not None:
            custom_cols.append((i, defn))

    visible_cache: dict = {}

    async def _visible_keys(kind: str) -> set:
        if kind not in visible_cache:
            subtype_defs = await custom_field_service.list_definitions_for_subtype(
                db, tenant_id, _ENTITY_TYPE, kind)
            visible_cache[kind] = {d.field_key for d in subtype_defs}
        return visible_cache[kind]

    created = updated = 0
    errors: list[ImportError] = []

    for rownum, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None for v in row):
            continue
        title = _cell(row, idx["title"])
        kind = _cell(row, idx["change_kind"])
        if not title:
            errors.append(ImportError(row=rownum, field="title", message="title is required"))
            continue
        if not kind:
            errors.append(ImportError(row=rownum, field="change_kind", message="change_kind is required"))
            continue

        custom_fields: dict = {}
        for col_idx, defn in custom_cols:
            raw = row[col_idx] if col_idx < len(row) else None
            if raw is None or (isinstance(raw, str) and raw.strip() == ""):
                continue
            custom_fields[defn.field_key] = _coerce(raw, defn.field_type)
        try:
            await custom_field_service.validate_custom_fields(
                db, tenant_id, _ENTITY_TYPE, custom_fields,
                visible_field_keys=await _visible_keys(kind),
            )
        except HTTPException as e:
            errors.append(ImportError(row=rownum, field="custom_fields", message=str(e.detail)))
            continue

        ek = _cell(row, idx["external_key"])
        existing = None
        if ek:
            existing = (await db.execute(
                select(ReleaseChange).where(
                    ReleaseChange.tenant_id == tenant_id,
                    ReleaseChange.release_id == release_id,
                    ReleaseChange.external_key == ek,
                    ReleaseChange.deleted_at.is_(None),
                )
            )).scalar_one_or_none()

        fields = dict(
            title=title, change_kind=kind,
            description=_cell(row, idx["description"]),
            external_status=_cell(row, idx["external_status"]),
            project_code=_cell(row, idx["project_code"]),
            project_name=_cell(row, idx["project_name"]),
            custom_fields=custom_fields or None,
        )
        if existing:
            for k, v in fields.items():
                setattr(existing, k, v)
            existing.source = "spreadsheet"
            updated += 1
        else:
            db.add(ReleaseChange(
                tenant_id=tenant_id, release_id=release_id, external_key=ek,
                source="spreadsheet", **fields,
            ))
            created += 1

    await db.flush()
    return {"created": created, "updated": updated, "errors": errors}
