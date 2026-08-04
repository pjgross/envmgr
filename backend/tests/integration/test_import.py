"""Integration tests for Excel Import (M5)."""
import io

import openpyxl
import pytest
import pytest_asyncio
from httpx import AsyncClient

from app.services.environment_tier_defaults import (
    seed_environment_tier_defaults_for_tenant,
)
from tests.factories import post_environment


@pytest_asyncio.fixture
async def seeded_tiers(db_session, test_tenant):
    """The tier vocabulary the import resolves against.

    Production tenants get it from tenant_service.create_tenant; the bare
    `test_tenant` fixture is built as a raw row, so it has none.
    """
    await seed_environment_tier_defaults_for_tenant(db_session, test_tenant.id)
    await db_session.commit()


# ---------------------------------------------------------------------------
# Excel file helpers
# ---------------------------------------------------------------------------


def make_environment_excel(rows: list[dict]) -> bytes:
    """Build an Excel workbook with environment rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Name", "Type", "Description"])
    for row in rows:
        ws.append([row.get("Name", ""), row.get("Type", ""), row.get("Description", "")])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_environment_excel_no_name_col(rows: list[dict]) -> bytes:
    """Build an Excel workbook WITHOUT a Name column — for error testing."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Type", "Description"])
    for row in rows:
        ws.append([row.get("Type", ""), row.get("Description", "")])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_system_excel(rows: list[dict]) -> bytes:
    """Build an Excel workbook with system rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Name", "Description", "GitHub URL"])
    for row in rows:
        ws.append([row.get("Name", ""), row.get("Description", ""), row.get("GitHub URL", "")])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Tests — Environments
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_import_environments(client: AsyncClient, auth_headers, seeded_tiers):
    """POST /import/environments with a valid file creates new environments."""
    file_bytes = make_environment_excel([
        {"Name": "ImportedEnv1", "Type": "uat", "Description": "First imported env"},
        {"Name": "ImportedEnv2", "Type": "staging"},
    ])

    resp = await client.post(
        "/api/v1/import/environments",
        headers=auth_headers,
        files={"file": ("envs.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["created"] == 2
    assert data["skipped"] == 0
    assert data["errors"] == []

    # Verify environments exist via list endpoint
    list_resp = await client.get("/api/v1/environments/", headers=auth_headers)
    names = [e["name"] for e in list_resp.json()]
    assert "ImportedEnv1" in names
    assert "ImportedEnv2" in names


@pytest.mark.asyncio
async def test_import_environments_skip_existing(
    client: AsyncClient, auth_headers, seeded_tiers
):
    """Importing an environment whose name already exists → skipped, not error."""
    # Pre-create the environment
    await post_environment(client, auth_headers, "ExistingEnv")

    file_bytes = make_environment_excel([
        {"Name": "ExistingEnv", "Type": "uat"},
        {"Name": "BrandNewEnv", "Type": "staging"},
    ])

    resp = await client.post(
        "/api/v1/import/environments",
        headers=auth_headers,
        files={"file": ("envs.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["created"] == 1
    assert data["skipped"] == 1
    assert data["errors"] == []


@pytest.mark.asyncio
async def test_import_sets_the_importing_admin_as_owner_with_no_expiry(
    client: AsyncClient, auth_headers, seeded_tiers
):
    """The importer is present, acting and identifiable, so recording them as
    owner is truthful — unlike fabricating an owner for a pre-existing row,
    which the spec deliberately refused. `expires_at` stays null: the
    spreadsheet has none to offer, and null now means "no expiry planned"."""
    file_bytes = make_environment_excel([
        {"Name": "ImportedWithOwner", "Type": "uat"},
    ])

    resp = await client.post(
        "/api/v1/import/environments",
        headers=auth_headers,
        files={"file": ("envs.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200, resp.text
    assert resp.json()["created"] == 1

    me = await client.get("/api/v1/auth/me", headers=auth_headers)

    list_resp = await client.get("/api/v1/environments/", headers=auth_headers)
    [env] = [e for e in list_resp.json() if e["name"] == "ImportedWithOwner"]
    assert env["owner_username"] == me.json()["username"]
    assert env["expires_at"] is None


@pytest.mark.asyncio
async def test_an_imported_environment_can_be_patched_without_supplying_an_expiry(
    client: AsyncClient, auth_headers, seeded_tiers
):
    """The defect this fix closes: before Decision 1, a null expiry counted as
    a governance gap and the PATCH compliance rule refused every patch until
    both owner and expiry were supplied — freezing every imported row,
    because the importer sets an owner but deliberately no expiry. A
    description-only patch must now succeed on its own."""
    file_bytes = make_environment_excel([
        {"Name": "ImportedThenPatched", "Type": "uat"},
    ])
    resp = await client.post(
        "/api/v1/import/environments",
        headers=auth_headers,
        files={"file": ("envs.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200, resp.text

    list_resp = await client.get("/api/v1/environments/", headers=auth_headers)
    [env] = [e for e in list_resp.json() if e["name"] == "ImportedThenPatched"]
    assert env["expires_at"] is None  # imported rows have no expiry, by design

    patched = await client.patch(
        f"/api/v1/environments/{env['id']}",
        headers=auth_headers,
        json={"description": "reviewed, still no expiry planned"},
    )
    assert patched.status_code == 200, patched.text
    assert patched.json()["expires_at"] is None


@pytest.mark.asyncio
async def test_import_by_an_impersonating_master_admin_imports_unowned_not_404(
    client: AsyncClient, db_session, test_tenant, seeded_tiers
):
    """Regression this branch introduced, visible only under impersonation.

    A master admin impersonating a tenant has `active_tenant_id` == the
    impersonated tenant, but their own `User.tenant_id` is the *system*
    tenant they actually belong to. Before this fix, the route passed
    `current_user.id` straight through as `owner_user_id` while scoping the
    import to `active_tenant_id` — under impersonation the two disagree, so
    `_validate_client_foreign_keys`'s `User.tenant_id == tenant_id` check 404'd
    ("Owner not found"). That `HTTPException` is not caught by the per-row
    `except (ValueError, ValidationError)` in excel_import_service, so the
    *entire upload* failed with a message naming neither the file nor the
    real cause. No existing test covered the impersonation dimension of the
    import path, so this shipped and stayed green.
    """
    from app.core.security import get_password_hash
    from app.db.models.user import Tenant, User

    system_tenant = Tenant(name="System Org", slug="system-org-import-regression")
    db_session.add(system_tenant)
    await db_session.commit()
    await db_session.refresh(system_tenant)

    master = User(
        tenant_id=system_tenant.id,
        username="masteradmin-import-regression",
        email="masteradmin-import-regression@test.com",
        password_hash=get_password_hash("masterpass1"),
        role="Admin",
        is_active=True,
        is_master_admin=True,
    )
    db_session.add(master)
    await db_session.commit()
    await db_session.refresh(master)

    login = await client.post("/api/v1/auth/login", json={
        "username": master.username,
        "password": "masterpass1",
        "tenant_slug": system_tenant.slug,
    })
    assert login.status_code == 200, login.text
    master_token = login.json()["access_token"]

    sign_in_as = await client.post(
        f"/api/v1/admin/tenants/{test_tenant.id}/sign-in-as",
        headers={"Authorization": f"Bearer {master_token}"},
    )
    assert sign_in_as.status_code == 200, sign_in_as.text
    imp_headers = {"Authorization": f"Bearer {sign_in_as.json()['access_token']}"}

    file_bytes = make_environment_excel([
        {"Name": "ImpersonatedImport", "Type": "uat"},
    ])

    resp = await client.post(
        "/api/v1/import/environments",
        headers=imp_headers,
        files={"file": ("envs.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    # The regression: this used to be a 404 ("Owner not found") for the
    # whole upload, not a 200 with one unowned row.
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["created"] == 1
    assert data["errors"] == []

    list_resp = await client.get("/api/v1/environments/", headers=imp_headers)
    [env] = [e for e in list_resp.json() if e["name"] == "ImpersonatedImport"]
    assert env["owner_user_id"] is None
    assert env["owner_username"] is None

    # Reported, not silently dropped — see the tier_fallbacks tests below.
    assert any(
        fb["field"] == "Owner" and fb["row"] == 2 for fb in data["tier_fallbacks"]
    )


@pytest.mark.asyncio
async def test_import_reports_a_type_fallback_in_the_summary(
    client: AsyncClient, auth_headers, seeded_tiers
):
    """The spec says a fallback to Other is 'reported in the import summary'.
    Before this fix it happened silently: the result dict was
    {created, skipped, errors} with no fallback channel at all, so an admin
    importing 200 rows with a mistyped Type column got everything filed
    under Other with no signal.

    The discriminating half is the second row: it matches a real tier, so it
    must NOT show up in tier_fallbacks — otherwise this channel could be
    "every created row" and the test wouldn't tell the difference.
    """
    file_bytes = make_environment_excel([
        {"Name": "MistypedType", "Type": "not-a-real-tier"},
        {"Name": "RealTypeMatch", "Type": "uat"},
    ])

    resp = await client.post(
        "/api/v1/import/environments",
        headers=auth_headers,
        files={"file": ("envs.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["created"] == 2

    fallbacks = data["tier_fallbacks"]
    assert [fb["row"] for fb in fallbacks if fb["field"] == "Type"] == [2]
    assert "not-a-real-tier" in fallbacks[0]["message"]
    # The matched-tier row is the discriminating half: it must not appear.
    assert all(fb["row"] != 3 for fb in fallbacks)


@pytest.mark.asyncio
async def test_import_reports_a_blank_type_fallback_distinctly(
    client: AsyncClient, auth_headers, seeded_tiers
):
    """A blank Type and a mistyped one are different facts for the admin to
    act on, so the message says which happened."""
    file_bytes = make_environment_excel([
        {"Name": "BlankType", "Type": ""},
    ])

    resp = await client.post(
        "/api/v1/import/environments",
        headers=auth_headers,
        files={"file": ("envs.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    [fb] = data["tier_fallbacks"]
    assert fb["field"] == "Type"
    assert "no type specified" in fb["message"].lower()


@pytest.mark.asyncio
async def test_import_environments_missing_required_col(client: AsyncClient, auth_headers):
    """File without Name column → HTTP 400."""
    file_bytes = make_environment_excel_no_name_col([
        {"Type": "uat", "Description": "No name column"},
    ])

    resp = await client.post(
        "/api/v1/import/environments",
        headers=auth_headers,
        files={"file": ("envs.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 400
    assert "Name" in resp.json()["detail"]


# ---------------------------------------------------------------------------
# Tests — Systems
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_import_systems(client: AsyncClient, auth_headers):
    """POST /import/systems with a valid file creates new systems."""
    file_bytes = make_system_excel([
        {"Name": "ImportedSystem1", "Description": "First system"},
        {"Name": "ImportedSystem2", "GitHub URL": "https://github.com/example/repo"},
    ])

    resp = await client.post(
        "/api/v1/import/systems",
        headers=auth_headers,
        files={"file": ("systems.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["created"] == 2
    assert data["skipped"] == 0
    assert data["errors"] == []

    # Verify systems exist via list endpoint
    list_resp = await client.get("/api/v1/systems/", headers=auth_headers)
    names = [s["name"] for s in list_resp.json()]
    assert "ImportedSystem1" in names
    assert "ImportedSystem2" in names


@pytest.mark.asyncio
async def test_import_systems_skip_existing(client: AsyncClient, auth_headers):
    """Importing a system whose name already exists → skipped, not error."""
    # Pre-create the system
    await client.post(
        "/api/v1/systems/",
        headers=auth_headers,
        json={"name": "ExistingSystem"},
    )

    file_bytes = make_system_excel([
        {"Name": "ExistingSystem"},
        {"Name": "FreshSystem"},
    ])

    resp = await client.post(
        "/api/v1/import/systems",
        headers=auth_headers,
        files={"file": ("systems.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["created"] == 1
    assert data["skipped"] == 1
    assert data["errors"] == []
